# ============================================================
# excel_export.py — Export Presensi ke File .xlsx
# ============================================================

import os
import re
import requests
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from config import SERVER, EXPORT_FOLDER


def _safe_filename(nama: str) -> str:
    """Bersihkan nama untuk dijadikan nama file."""
    return re.sub(r'[\\/*?:"<>|]', "", nama).strip()


def export_presensi(acara_id: int, nama_acara: str, tanggal: str) -> str:
    """
    Mengambil data presensi dari server, lalu mengekspor ke file Excel.
    Returns: path file .xlsx yang dibuat, atau raise Exception jika gagal.
    """
    if not OPENPYXL_OK:
        raise ImportError(
            "Library 'openpyxl' belum terinstall.\n"
            "Jalankan: pip install openpyxl"
        )

    # ── 1. Ambil data presensi dari server ─────────────────────
    url  = f"{SERVER}/presensi.php?action=list&acara_id={acara_id}"
    resp = requests.get(url, timeout=10)
    resp.raise_for_status()
    payload = resp.json()

    if payload.get("status") != "ok":
        raise Exception(payload.get("pesan", "Gagal mengambil data"))

    presensi_data = payload.get("data", [])

    # ── 2. Ambil statistik ─────────────────────────────────────
    url_stat = f"{SERVER}/presensi.php?action=statistik&acara_id={acara_id}"
    resp_stat = requests.get(url_stat, timeout=10)
    stat = resp_stat.json() if resp_stat.ok else {}

    # ── 3. Ambil semua mahasiswa (untuk tidak hadir) ───────────
    url_mhs  = f"{SERVER}/mahasiswa.php?action=list"
    resp_mhs = requests.get(url_mhs, timeout=10)
    semua_mhs = resp_mhs.json().get("data", []) if resp_mhs.ok else []

    # Buat set mahasiswa yang hadir
    hadir_ids = {str(p.get("nim", "")) for p in presensi_data}

    # ── 4. Buat Workbook ────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presensi"

    # ── Warna & Style ──────────────────────────────────────────
    C_HEADER   = "6C63FF"   # ungu header
    C_HADIR    = "D4EDDA"   # hijau muda
    C_TERLAMBAT = "FFF3CD"  # kuning muda
    C_TIDAK    = "F8D7DA"   # merah muda
    C_TITLE    = "0F1117"   # hitam
    C_ALT      = "F5F5FB"   # abu abu

    thin = Side(border_style="thin", color="CCCCCC")
    thick = Side(border_style="medium", color="9999AA")
    border_thin  = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    def hfill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def hfont(bold=False, color="000000", size=11):
        return Font(bold=bold, color=color, size=size, name="Arial")

    def halign(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── Row 1: Judul Besar ─────────────────────────────────────
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value         = "DAFTAR PRESENSI MAHASISWA"
    c.font          = hfont(bold=True, color="FFFFFF", size=16)
    c.fill          = hfill(C_HEADER)
    c.alignment     = halign()

    # ── Row 2: Info Acara ──────────────────────────────────────
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = f"Acara: {nama_acara}  |  Tanggal: {tanggal}  |  Dicetak: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font      = hfont(size=10, color="444466")
    c.fill      = hfill("EDEDF8")
    c.alignment = halign()

    # ── Row 3: Kosong ──────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: Header Kolom ────────────────────────────────────
    headers = ["No", "NIM", "Nama Lengkap", "Program Studi", "Angkatan",
               "Waktu Tap", "Status Hadir", "Keterangan"]
    ws.row_dimensions[4].height = 22

    for col_i, h in enumerate(headers, start=1):
        cell            = ws.cell(row=4, column=col_i, value=h)
        cell.font       = hfont(bold=True, color="FFFFFF", size=11)
        cell.fill       = hfill(C_HEADER)
        cell.alignment  = halign()
        cell.border     = border_thin

    # ── Row 5+: Data Hadir ─────────────────────────────────────
    row_num = 5
    no      = 1

    for p in presensi_data:
        ws.row_dimensions[row_num].height = 18
        status    = p.get("status_hadir", "hadir")
        row_color = C_HADIR if status == "hadir" else C_TERLAMBAT
        waktu_raw = p.get("waktu_tap", "")
        try:
            waktu_fmt = datetime.strptime(waktu_raw, "%Y-%m-%d %H:%M:%S").strftime("%H:%M:%S")
        except Exception:
            waktu_fmt = waktu_raw

        ket = "Tepat Waktu" if status == "hadir" else "Terlambat"

        row_data = [
            no,
            p.get("nim", ""),
            p.get("nama", ""),
            p.get("prodi", ""),
            p.get("angkatan", ""),
            waktu_fmt,
            status.upper(),
            ket
        ]

        for col_i, val in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_num, column=col_i, value=val)
            cell.fill      = hfill(row_color)
            cell.border    = border_thin
            cell.alignment = halign(h="center" if col_i in [1, 5, 6, 7] else "left")
            cell.font      = hfont(size=10)

        row_num += 1
        no      += 1

    # ── Data Tidak Hadir (mahasiswa tanpa presensi) ────────────
    tidak_hadir_list = [m for m in semua_mhs if str(m.get("nim", "")) not in hadir_ids]

    for m in tidak_hadir_list:
        ws.row_dimensions[row_num].height = 18
        row_data = [
            no,
            m.get("nim", ""),
            m.get("nama", ""),
            m.get("prodi", ""),
            m.get("angkatan", ""),
            "-",
            "TIDAK HADIR",
            "-"
        ]
        for col_i, val in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_num, column=col_i, value=val)
            cell.fill      = hfill(C_TIDAK)
            cell.border    = border_thin
            cell.alignment = halign(h="center" if col_i in [1, 5, 6, 7] else "left")
            cell.font      = hfont(size=10)

        row_num += 1
        no      += 1

    # ── Baris Kosong ───────────────────────────────────────────
    row_num += 1

    # ── Summary ────────────────────────────────────────────────
    total_hadir  = stat.get("total_hadir", len(presensi_data))
    total_mhs    = stat.get("total_mhs", len(semua_mhs))
    terlambat    = stat.get("terlambat", 0)
    tidak_hadir  = total_mhs - total_hadir

    summary_rows = [
        ("Total Mahasiswa",  total_mhs,   "222244"),
        ("Total Hadir",      total_hadir, "155724"),
        ("Terlambat",        terlambat,   "856404"),
        ("Tidak Hadir",      tidak_hadir, "842029"),
    ]

    ws.merge_cells(f"A{row_num}:F{row_num}")
    hdr = ws.cell(row=row_num, column=1, value="REKAPITULASI")
    hdr.font      = hfont(bold=True, color="FFFFFF", size=12)
    hdr.fill      = hfill(C_HEADER)
    hdr.alignment = halign()

    row_num += 1

    for label, value, txt_color in summary_rows:
        ws.merge_cells(f"A{row_num}:E{row_num}")
        lbl_cell           = ws.cell(row=row_num, column=1, value=label)
        lbl_cell.font      = hfont(bold=True, color=txt_color, size=11)
        lbl_cell.alignment = halign(h="right")
        lbl_cell.border    = border_thin

        val_cell           = ws.cell(row=row_num, column=6, value=value)
        val_cell.font      = hfont(bold=True, color=txt_color, size=11)
        val_cell.alignment = halign()
        val_cell.border    = border_thin
        row_num += 1

    # ── Lebar Kolom ────────────────────────────────────────────
    col_widths = [5, 18, 30, 28, 10, 14, 14, 15]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze header ──────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Simpan file ────────────────────────────────────────────
    safe_nama   = _safe_filename(nama_acara)
    tgl_str     = tanggal.replace("-", "")
    filename    = f"Presensi_{safe_nama}_{tgl_str}.xlsx"
    filepath    = os.path.join(EXPORT_FOLDER, filename)

    wb.save(filepath)
    return filepath
